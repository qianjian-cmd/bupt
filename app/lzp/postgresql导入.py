# -*- coding:utf-8 -*- 
import os
import asyncpg
import openpyxl
import csv
from fastapi import FastAPI, Query, File, UploadFile
import uvicorn
import uuid
import xlrd
import xlwt
import sys
import asyncio
from win32com.client import Dispatch

fileDict={}                     #文件字典
standardTable=["tbCell","tbKPI","tbPRB","tbMROData"]
taskList=[]
primaryDict={'tbCell':['SECTOR_ID'],'tbKPI':['小区名称'],'tbPRB':['起始时间','小区名'],'tbMROData':['TimeStamp','ServingSector','InterferingSector']}        #主键字典
primaryPlace={'tbCell': [1],'tbKPI': [3],'tbPRB':[0,3],'tbMROData':[0,1,2]}               #主键位置
typeDict={'tbCell':[0,0,0,1,0,1,1,1,1,1,0,2,2,0,0,2,2,2,2,2]}
typeDict['tbKPI']=[0,0,0,0,1,1,2,1,1,2,1,1,2,2,1,1,1,2,1,1,1,1,1,1,1,1,2,2,2,2,2,1,1,1,2,1,1,1,1,1,1]
typeDict['tbPRB']=[0,0,0,0]
for i in range(0,100):
    typeDict['tbPRB'].append(1)
typeDict['tbMROData']=[0,0,0,2,2,1,1]
#0-str、1-int、2-float
attDict={}
attDict['tbCell']=["CITY","SECTOR_ID" ,"SECTOR_NAME","ENODEBID","ENODEB_NAME","EARFCN","PCI","PSS","SSS","TAC","VENDOR","LONGITUDE","LATITUDE","STYLE","AZIMUTH","HEIGHT","ELECTTILT","MECHTILT","TOTLETILT"]
attDict['tbKPI']=["起始时间"  ,"网元基站名称"  ,"小区"  ,"小区名称"  ,"RRC连接建立完成次数"  ,"RRC连接请求次数"  ,"RRC建立成功率qf" ,"ERAB建立成功总次数" ,"ERAB建立尝试总次数" ,"ERAB建立成功率2"  ,"eNodeB触发的ERAB异常释放总次数" ,"小区切换出ERAB异常释放总次数" ,"ERAB掉线率" ,"无线接通率ay"  ,"eNodeB发起的S1RESET导致的UEContext释放次数"  ,"UEContext异常释放次数" ,"UEContext建立成功总次数" ,"无线掉线率"  ,"eNodeB内异频切换出成功次数" , "eNodeB内异频切换出尝试次数" ,"eNodeB内同频切换出成功次数"  ,"eNodeB内同频切换出尝试次数"  ,"eNodeB间异频切换出成功次数"  ,"eNodeB间异频切换出尝试次数"  ,"eNodeB间同频切换出成功次数"  ,"eNodeB间同频切换出尝试次数"  ,"eNB内切换成功率" ,"eNB间切换成功率"  ,"同频切换成功率zsp"  ,"异频切换成功率zsp"  ,"切换成功率"  ,"小区PDCP层所接收到的上行数据的总吞吐量"  ,"小区PDCP层所发送的下行数据的总吞吐量"  ,"RRC重建请求次数"  ,"RRC连接重建比率"  ,"通过重建回源小区的eNodeB间同频切换出执行成功次数" ,"通过重建回源小区的eNodeB间异频切换出执行成功次数" ,"通过重建回源小区的eNodeB内同频切换出执行成功次数"  ,"通过重建回源小区的eNodeB内异频切换出执行成功次数" ,"eNB内切换出成功次数" ,"eNB内切换出请求次数"]
attDict['tbPRB']=["起始时间","网元基站名称","小区描述","小区名"]
for i in range(0,100):
    sql="第"
    sql+=str(i)
    sql+="个PRB上检测到的干扰噪声的平均值"
    attDict['tbPRB'].append(sql)
attDict['tbMROData']=["TimeStamp","ServingSector","InterferingSector","LteScRSRP","LteNcRSRP","LteNcEarfcn","LteNcPci"]

taskRateDict={}
taskLast={}

app = FastAPI()


@app.post("/file/upload/")
async def upload(file: UploadFile = File(...)):
    global fileDict
    try:
        contents=await file.read()
        fileID=uuid.uuid4()
        path="./data/"
        path+=str(fileID)
        type=file.filename.split('.')
        path+="."
        path+=str(type[1])
        with open(path,'wb') as f:
            f.write(contents)
        fileName=str(fileID)
        fileName+='.'
        fileName+=str(type[1])
        fileDict[str(fileID)]=str(fileName)
        #print(fileDict[fileID])
        return {
                 "error": "ok",
                 "message": "string",
                 "data": {
                            "fileID": fileID,
                            #"fileName": fileName
                         }
               }
    except:
        return {
                 "error": "upload-failed",
                 "message": "string"
               }


@app.get("/db/import/create-task/")
async def create_task(
    *,
    name: str,
    fileID: str
):
    global standardTable          #任务所需的四个表
    global table                  #表名
    global fileDict               #文件字典
    global taskRateDict           #本次任务字典
    global taskLast           #上次任务字典
    try:
        taskID=uuid.uuid4()
        file_path="./data/"
        file_path+=str(fileDict[fileID])
        #print(file_path)
        get=0

        for i in range(0,4):
            if str(name)==standardTable[i]:
                get=1

        if get!=1:           #传来的表名不属于要求的四种
            return {
                     "error": "invalid-table-name",
                     "message": "string"
                   }
            
        else:                 #表名符合要求
            table=str(name)

            coro = openFile(file_path,taskID)          #异步并发
            loop = asyncio.get_event_loop()
            task = loop.create_task(coro)

            taskRateDict[str(taskID)]=float(0)
            taskLast[str(taskID)]=float(0)

            return {
                     "error": "ok",
                     "message": "string",
                     "data": {
                               "taskID": taskID
                             }

                   }

    except:                   #文件ID不存在
        return {
                 "error": "invalid-file-id",
                 "message": "string"
               }



@app.get("/db/import/task-status/")
async def task_status(
    *,
    taskID: str
):
    global taskRateDict           #本次任务字典
    global taskLast               #上次任务字典

    try:
        if taskRateDict[taskID]>=100:
            return{ "error": "ok",
                    "message": "string",
                    "data": {
                               "status": "success"
                            }
                  }
        else:
            while taskRateDict[taskID]==taskLast[taskID]:
                await asyncio.sleep(0.1)
            if taskRateDict[taskID]>=100:
                taskRateDict[taskID]=100
            taskLast[taskID]=taskRateDict[taskID]
            return{ "error": "ok",
                    "message": "string",
                    "data": {
                               "status": str(taskRateDict[taskID])
                            }
                  }

    except:
        return{ "error": "invalid-task-id",
                "message": "string",
              }


def just_open(filename):
    xlApp = Dispatch("Excel.Application")
    xlApp.Visible = False
    filePath=str(os.getcwd())
    mid=filename.split('.')
    filePath+=str(mid[1])
    filePath+="."
    filePath+=mid[2]
    #print(filePath)
    xlBook = xlApp.Workbooks.Open(filePath)
    xlBook.Save()
    xlBook.Close()


async def openFile(file_path,taskID):
    global row_number
    global col_number
    global dataResult
    global table

    #print(file_path)
    type=file_path.split('.')
    dataResult=[]
    #print(type[2])

    if type[2]=="xls":
        file = xlrd.open_workbook(file_path)       #打开Excel文件
        RSheet = file.sheet_by_index(0)            #根据sheet页的排序选取sheet
        row_number = RSheet.nrows #获取有数据的最大行数
        col_number = RSheet.ncols #获取有数据的最大列数
        for i in range(0,row_number):
            dataResult.append(RSheet.row_values(i))

    elif type[2]=="xlsx":
        if table=="tbCell":
            just_open(file_path)
        #print(1)
        workbook=openpyxl.load_workbook(file_path,data_only=True)
        #print(2)
        RSheet = workbook.worksheets[0]
        #print(3)
        row_number =RSheet.max_row
        col_number =RSheet.max_column
        for row in RSheet.rows:
            dataRow=[]
            for cell in row:
                dataRow.append(cell.value)
            dataResult.append(dataRow)

    elif type[2]=="csv":
        with open(file_path) as f:
            reader = csv.reader(f)
            dataResult=[row for row in reader]
            row_number = len(dataResult)                  #获取有数据的最大行数
            col_number = len(dataResult[0])                  #获取有数据的最大列数
            

    await read_data(taskID)                 #操作数据库


#从本地定量读取文件并存到数据库
async def read_data(taskID):
    global table            #表名
    global attDict          #属性字典
    global taskRateDict     #任务字典
    global row_content
    global col_content      #属性名
    global col_number
    global row_number
    global dataResult
    global typeDict
    global packet
    packet=100
    
    #row_content = dataResult[0]
    #print(len(row_content))
    row_content = attDict[table]
    #print(len(row_content))
    col_content=row_content
    #print(len(typeDict[table]))
    await creat_table()       #创建表项

    if (row_number-1)%packet==0:
        packetNum=int((row_number-1)/packet)
    else:
        packetNum=int((row_number-1)/packet)+1

    num=0              #目前处理到第几组
    for i in range(0,packetNum):          #分批(packet)读取
        await data_handle(num)            #数据处理(插入/更新)
        #print(i+1)
        num=num+1
        taskRateDict[str(taskID)]=(num*packet*100)/(row_number-1)
    #print("sucessful!")


#建表
async def creat_table():              
    global col_content
    global col_number
    global table

    try:
        con = await asyncpg.connect(database="grade", user="qian", password="230108", host="127.0.0.1", port="5432")
        print("连接成功")
    except:
        print(f'{e}')
        print('连接失败')
    
    if table=="tbCell":
        sql="create table if not exists tbCell (CITY text,SECTOR_ID text primary key,SECTOR_NAME text NOT NULL,ENODEBID int NOT NULL,ENODEB_NAME text NOT NULL,EARFCN int NOT NULL,PCI int,PSS int,SSS int,TAC int,VENDOR text,LONGITUDE float NOT NULL,LATITUDE float NOT NULL,STYLE text,AZIMUTH float NOT NULL,HEIGHT float,ELECTTILT float,MECHTILT float,TOTLETILT float)"
    elif table=="tbKPI":
        sql="create table if not exists tbKPI (起始时间 text ,网元基站名称 text ,小区 text ,小区名称 text ,RRC连接建立完成次数 int ,RRC连接请求次数 int ,RRC建立成功率qf float,ERAB建立成功总次数 int,ERAB建立尝试总次数 int,ERAB建立成功率2 float ,eNodeB触发的ERAB异常释放总次数 int,小区切换出ERAB异常释放总次数 int,ERAB掉线率 float,无线接通率ay float ,eNodeB发起的S1RESET导致的UEContext释放次数 int ,UEContext异常释放次数 int ,UEContext建立成功总次数 int ,无线掉线率 float ,eNodeB内异频切换出成功次数 int , eNodeB内异频切换出尝试次数 int,eNodeB内同频切换出成功次数 int ,eNodeB内同频切换出尝试次数 int ,eNodeB间异频切换出成功次数 int ,eNodeB间异频切换出尝试次数 int ,eNodeB间同频切换出成功次数 int ,eNodeB间同频切换出尝试次数 int ,eNB内切换成功率 float ,eNB间切换成功率 float ,同频切换成功率zsp float ,异频切换成功率zsp float ,切换成功率 float ,小区PDCP层所接收到的上行数据的总吞吐量 bigint ,小区PDCP层所发送的下行数据的总吞吐量 bigint ,RRC重建请求次数 int ,RRC连接重建比率 float ,通过重建回源小区的eNodeB间同频切换出执行成功次数 int ,通过重建回源小区的eNodeB间异频切换出执行成功次数 int,通过重建回源小区的eNodeB内同频切换出执行成功次数 int ,通过重建回源小区的eNodeB内异频切换出执行成功次数 int ,eNB内切换出成功次数 int ,eNB内切换出请求次数 int)"
    elif table=="tbPRB":
        sql="create table if not exists tbPRB (起始时间 text,网元基站名称 text,小区描述 text,小区名 text"
        for i in range(0,100):
            sql+=",第"
            sql+=str(i)
            sql+="个PRB上检测到的干扰噪声的平均值 int"
        sql+=",primary key(起始时间,小区名))"
    elif table=="tbMROData":
        sql="create table if not exists tbMROData (TimeStamp text,ServingSector text,InterferingSector text,LteScRSRP float,LteNcRSRP float,LteNcEarfcn int,LteNcPci int,PRIMARY KEY(TimeStamp,ServingSector,InterferingSector))"

    try:
        await con.fetch(sql)
    except Exception as e:
        print(f'{e}')
        print('创建表失败')

    await con.close()


#数据处理分类(插入/更新)
async def data_handle(num):
    global col_content
    global row_number
    global packet
    global dataResult
    global primaryDict      #主键字典
    global primaryPlace     #主键位置字典
    global typeDict         #类型字典
    global attDict          #属性字典
    global table

    con = await asyncpg.connect(database="grade", user="qian", password="230108", host="127.0.0.1", port="5432")

    if row_number-1-packet*(num+1)>=0:       #计算本次处理的数据量
        end=packet+1
    else:
        end=row_number-packet*num
        #print(end)

    insert_data=[]              #待插入的数据
    update_data=[]              #待更新的数据
    primary_data=[]             #主键数据
    try:
        for i in range(1,end):
            now_content=[]              #待处理的数据
            mid_content=dataResult[packet*num+i]
            error=0            
            primary_data.append([])
            for j in range(0,col_number):                 #消除空值
                #print(type(mid_content[j]))
                #if mid_content[j]=='':
                #    mid_content[j]="NULL"
                if typeDict[table][j]==1:          #类型转换
                    try:    
                        now_content.append(int(mid_content[j]))
                    except:                        #非法类型数据处理
                        now_content.append(None)
                
                elif typeDict[table][j]==2:          #类型转换
                    try:    
                        now_content.append(float(mid_content[j]))
                    except:                        #非法类型数据处理
                        now_content.append(None)                   
                else:
                    now_content.append(mid_content[j])

                if j in primaryPlace[table]:            #构建主键list
                    #print(j)
                    primary_data[i-1].append(now_content[j])

            sql="select * from "
            sql+=table
            sql+=" where "
            for k in range(0,len(primaryPlace[table])):
                 sql+=primaryDict[table][k] 
                 sql+=" = $"
                 sql+=str(k+1)
                 if k!=len(primaryPlace[table])-1:
                     sql+=" and "
            #print(sql)
            
            result=[]
            if len(primary_data[i-1])==1:
                result = await con.fetch(sql,primary_data[i-1][0])
            elif len(primary_data[i-1])==2:
                result = await con.fetch(sql,primary_data[i-1][0],primary_data[i-1][1])
            else:
                result = await con.fetch(sql,primary_data[i-1][0],primary_data[i-1][1],primary_data[i-1][2])
            #print(str(primary_data[i-1]))
            #print(result)
            
            if result!='' and result!=None:
                update_data.append(now_content)            #构建更新队列 
                #print(now_content)
            else:
                insert_data.append(now_content)            #构建插入队列
                #print(now_content)
    except Exception as e:
        print(f'{e}')
        print('查询失败')
   
    await con.close()

    #print("update:")
    #print(update_data)
    #print("insert:")
    #print(insert_data)
    await update_table(update_data)                #批量更新数据
    await insert_table(insert_data)                #批量插入数据



#插入数据
async def insert_table(insert_data): 
    global col_content
    global table

    con = await asyncpg.connect(database="grade", user="qian", password="230108", host="127.0.0.1", port="5432")

    sql = "insert into "
    sql+=table
    sql+="("
    sql+=str(col_content[0])
    j=1
    for i in range(1,col_number):
        sql+=","
        sql+=str(col_content[i])
    sql+=") values("
    sql+="$"
    sql+=str(j)
    j=j+1
    for i in range(1,col_number):
        sql+=","
        sql+="$"
        sql+=str(j)
        j=j+1
            
    sql+=')'
    #print(sql)
    #print(insert_data)

    try:
        await con.executemany(sql,insert_data)
    except Exception as e:
        print(f'{e}')
        print('插入失败')
   
    await con.close()


#更新数据
async def update_table(update_data): 
    global col_content
    global table
    global primaryDict        #主键字典
    global primaryPlace       #主键位置

    con = await asyncpg.connect(database="grade", user="qian", password="230108", host="127.0.0.1", port="5432")

    sql = "update "
    sql+=table
    sql+=" set "
    sql+=str(col_content[0])
    j=1
    for i in range(1,col_number):
        sql+="= $"
        sql+=str(j)
        j=j+1
        sql+=","
        sql+=str(col_content[i])
    sql+="= $"
    sql+=str(j)
    j=j+1
    sql+=" where "

    for k in range(0,len(primaryPlace[table])):
         sql+=primaryDict[table][k] 
         sql+=" = $"
         sql+=str(j)
         j=j+1
         if k!=len(primaryPlace[table])-1:
             sql+=" and "
    #print(sql)
    #print(update_data[0][0])
    
    for i in range(0,len(update_data)):            #调整update属性的顺序
        for j in range(0,len(col_content)):
            if j in primaryPlace[table]:
                update_data[i].append(update_data[i][j])

    try:
        #print(update_data)
        await con.executemany(sql,update_data)
    except Exception as e:
        print(f'{e}')
        print('更新失败')
   
    await con.close()


#async def dataClean(now_content):          #数据清洗
#    pass


if __name__ == '__main__':
    uvicorn.run(app=app,host="127.0.0.1",port=8000)


